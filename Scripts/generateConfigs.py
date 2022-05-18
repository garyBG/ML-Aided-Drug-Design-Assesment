import openpyxl

excelPath = "Binding_Sites.xlsx"
tempPath = "template.file"
worksheet = openpyxl.open("Binding_Sites.xlsx", read_only=True).worksheets[0]

templateFile = open(tempPath).read()

for x in range(2, worksheet.max_row + 1):
    newConfig = templateFile
    newConfig = newConfig.replace("<ID>", str(worksheet.cell(x, 1).value))
    newConfig = newConfig.replace("<CENTER>", str(worksheet.cell(x, 2).value) + ", " + str(worksheet.cell(x, 3).value) + ", " + str(worksheet.cell(x, 4).value))
    newFile = open("configs\\" + str(worksheet.cell(x, 1).value) + ".file", "w")
    newFile.write(newConfig)
    newFile.close()
